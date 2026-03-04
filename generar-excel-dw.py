#!/usr/bin/env python3
"""
Genera un Excel con datos del DW para crear gráficos manualmente.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Estilos
header_font = Font(name='Calibri', bold=True, size=11)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
num_fmt = '#,##0'

def write_table(ws, headers, data, start_row=1):
    """Escribe tabla con formato."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for r, row in enumerate(data, start_row + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal='center')
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ''))
            for r in range(start_row, start_row + len(data) + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 4, 12)

# ============================================================================
# HOJA 1: Pacientes por Sucursal
# ============================================================================
ws1 = wb.active
ws1.title = 'Pacientes por Sucursal'
write_table(ws1,
    ['Sucursal', 'Total Pacientes'],
    [
        ['Grupo 1', 87487],
        ['Grupo 6', 34500],
        ['Grupo 3', 4500],
    ]
)
# Total
ws1.cell(row=5, column=1, value='TOTAL').font = Font(bold=True)
ws1.cell(row=5, column=1).border = thin_border
ws1.cell(row=5, column=2, value=126487).font = Font(bold=True)
ws1.cell(row=5, column=2).number_format = num_fmt
ws1.cell(row=5, column=2).border = thin_border
ws1.cell(row=5, column=2).alignment = Alignment(horizontal='center')

# ============================================================================
# HOJA 2: Atenciones por Sucursal
# ============================================================================
ws2 = wb.create_sheet('Atenciones por Sucursal')
write_table(ws2,
    ['Sucursal', 'Total Atenciones'],
    [
        ['Grupo 1', 68123],
        ['Grupo 6', 40134],
        ['Grupo 3', 15000],
    ]
)
ws2.cell(row=5, column=1, value='TOTAL').font = Font(bold=True)
ws2.cell(row=5, column=1).border = thin_border
ws2.cell(row=5, column=2, value=123257).font = Font(bold=True)
ws2.cell(row=5, column=2).number_format = num_fmt
ws2.cell(row=5, column=2).border = thin_border
ws2.cell(row=5, column=2).alignment = Alignment(horizontal='center')

# ============================================================================
# HOJA 3: Médicos por Sucursal
# ============================================================================
ws3 = wb.create_sheet('Medicos por Sucursal')
write_table(ws3,
    ['Sucursal', 'Total Médicos'],
    [
        ['Grupo 1', 1100],
        ['Grupo 6', 831],
        ['Grupo 3', 500],
    ]
)
ws3.cell(row=5, column=1, value='TOTAL').font = Font(bold=True)
ws3.cell(row=5, column=1).border = thin_border
ws3.cell(row=5, column=2, value=2431).font = Font(bold=True)
ws3.cell(row=5, column=2).number_format = num_fmt
ws3.cell(row=5, column=2).border = thin_border
ws3.cell(row=5, column=2).alignment = Alignment(horizontal='center')

# ============================================================================
# HOJA 4: Resumen General
# ============================================================================
ws4 = wb.create_sheet('Resumen General')
write_table(ws4,
    ['Sucursal', 'Host', 'Pacientes', 'Médicos', 'Atenciones'],
    [
        ['Grupo 3', 'localhost:5432 (PostgreSQL)', 4500, 500, 15000],
        ['Grupo 1', 'Supabase (aws-us-west-2)', 87487, 1100, 68123],
        ['Grupo 6', 'PostgreSQL 17 (dump)', 34500, 831, 40134],
    ]
)
ws4.cell(row=5, column=1, value='TOTAL').font = Font(bold=True)
ws4.cell(row=5, column=1).border = thin_border
for col, val in enumerate([None, 126487, 2431, 123257], 2):
    cell = ws4.cell(row=5, column=col, value=val)
    cell.font = Font(bold=True)
    cell.border = thin_border
    if val:
        cell.number_format = num_fmt
        cell.alignment = Alignment(horizontal='center')

out = '/home/tunek/Universidad/MATERIAS/bd-clinica/datos_dw_graficos.xlsx'
wb.save(out)
print(f'Excel generado: {out}')
