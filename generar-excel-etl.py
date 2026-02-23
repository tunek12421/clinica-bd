#!/usr/bin/env python3
"""
Genera Excel con datos y gráficos del ETL completo (G1 + G6):
- Hoja 1: Citas por Especialidad (Antes / Después)
- Hoja 2: Citas por Mes (Antes / Después)
- Hoja 3: Diagnósticos por Tipo (Antes / Después)
"""

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUTPUT = '/home/tunek/Universidad/MATERIAS/bd-clinica/graficas_etl_grupo1.xlsx'
IMG_DIR = '/tmp/etl_excel_charts2'
os.makedirs(IMG_DIR, exist_ok=True)

# =============================================================================
# DATOS ACTUALIZADOS (G3 + G1 + G6)
# =============================================================================
esp_antes = [
    ('Medicina General', 3542), ('Cardiología', 3097), ('Pediatría', 2170),
    ('Dermatología', 1905), ('Neurología', 1764), ('Ginecología', 1660),
    ('Oftalmología', 1220), ('Traumatología', 1001), ('Otorrinolaringología', 760),
    ('Urología', 741), ('Gastroenterología', 706), ('Neumología', 483),
    ('Psiquiatría', 426), ('Endocrinología', 331), ('Oncología', 194),
]
esp_despues = [
    ('Medicina General', 71813), ('Traumatología', 30307), ('Urología', 20429),
    ('Cardiología', 18921), ('Pediatría', 17088), ('Neurología', 16830),
    ('Dermatología', 16157), ('Ginecología', 14473), ('Oftalmología', 10706),
    ('Oncología', 10229), ('Otorrinolaringología', 8756), ('Endocrinología', 6514),
    ('Psiquiatría', 6236), ('Gastroenterología', 706), ('Neumología', 483),
]

meses = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
mes_antes = [1432, 1432, 1707, 1682, 1780, 2406, 2573, 2660, 1230, 1178, 1198, 722]
mes_despues = [23268, 21466, 23532, 22924, 22545, 22796, 23362, 23357, 21417, 22013, 21623, 21697]

diag_antes = [
    ('Clínico', 2142), ('Por Imagen', 1735), ('Laboratorio', 1389),
    ('Diferencial', 1228), ('Presuntivo', 1084), ('Definitivo', 1027),
    ('Prenatal', 890), ('Molecular', 750), ('Patológico', 737), ('Funcional', 620),
]
diag_despues = [
    ('Definitivo', 43780), ('Presuntivo', 43721), ('Clínico', 33940),
    ('Laboratorio', 15025), ('Ambulatorio', 8954), ('Por Imagen', 6307),
    ('Nutricional', 4802), ('Endoscópico', 4628), ('Diferencial', 1228), ('Prenatal', 890),
]

# =============================================================================
# ESTILOS
# =============================================================================
hf = Font(bold=True, color='FFFFFF', size=11)
fill_a = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
fill_d = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
tf = Font(bold=True, size=13, color='1F3A5F')
border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def style_h(ws, row, c1, c2, fill):
    for c in range(c1, c2+1):
        cl = ws.cell(row=row, column=c)
        cl.font = hf; cl.fill = fill; cl.alignment = Alignment(horizontal='center'); cl.border = border

def style_d(ws, row, c1, c2):
    for c in range(c1, c2+1):
        cl = ws.cell(row=row, column=c); cl.border = border
        if c > c1: cl.number_format = '#,##0'; cl.alignment = Alignment(horizontal='right')

def color_list(n, cmap='tab20'):
    cm = plt.get_cmap(cmap)
    return [cm(i/n) for i in range(n)]

# =============================================================================
# GRÁFICOS
# =============================================================================
plt.rcParams['font.size'] = 10; plt.rcParams['figure.dpi'] = 150

# Especialidades comparativo
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))
na = [e[0] for e in esp_antes]; va = [e[1] for e in esp_antes]
ax1.barh(na[::-1], va[::-1], color=color_list(len(na))[::-1])
ax1.set_title('ANTES (Solo Grupo 3)', fontweight='bold', color='#4472C4')
for i, v in enumerate(va[::-1]): ax1.text(v+50, i, f'{v:,}', va='center', fontsize=8)
nd = [e[0] for e in esp_despues]; vd = [e[1] for e in esp_despues]
ax2.barh(nd[::-1], vd[::-1], color=color_list(len(nd))[::-1])
ax2.set_title('DESPUÉS (G3 + G1 + G6)', fontweight='bold', color='#ED7D31')
for i, v in enumerate(vd[::-1]): ax2.text(v+500, i, f'{v:,}', va='center', fontsize=8)
fig.suptitle('Citas por Especialidad — Antes y Después del ETL', fontsize=14, fontweight='bold')
plt.tight_layout(); plt.savefig(f'{IMG_DIR}/esp.png', bbox_inches='tight'); plt.close()

# Meses comparativo
fig, ax = plt.subplots(figsize=(14, 6))
x = np.arange(12); w = 0.35
b1 = ax.bar(x-w/2, mes_antes, w, label='Antes (Grupo 3)', color='#4472C4')
b2 = ax.bar(x+w/2, mes_despues, w, label='Después (G3+G1+G6)', color='#ED7D31')
ax.set_xticks(x); ax.set_xticklabels(meses); ax.set_ylabel('Total de Citas')
ax.set_title('Citas por Mes — Antes y Después del ETL', fontweight='bold', fontsize=14); ax.legend()
for b in b1: ax.text(b.get_x()+b.get_width()/2, b.get_height()+200, f'{int(b.get_height()):,}', ha='center', fontsize=7, color='#4472C4')
for b in b2: ax.text(b.get_x()+b.get_width()/2, b.get_height()+200, f'{int(b.get_height()):,}', ha='center', fontsize=7, color='#ED7D31')
plt.tight_layout(); plt.savefig(f'{IMG_DIR}/mes.png', bbox_inches='tight'); plt.close()

# Diagnósticos comparativo
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
na = [d[0] for d in diag_antes]; va = [d[1] for d in diag_antes]
ax1.barh(na[::-1], va[::-1], color=color_list(len(na), 'Set3')[::-1])
ax1.set_title('ANTES (Solo Grupo 3)', fontweight='bold', color='#4472C4')
for i, v in enumerate(va[::-1]): ax1.text(v+20, i, f'{v:,}', va='center', fontsize=8)
nd = [d[0] for d in diag_despues]; vd = [d[1] for d in diag_despues]
ax2.barh(nd[::-1], vd[::-1], color=color_list(len(nd), 'Set3')[::-1])
ax2.set_title('DESPUÉS (G3 + G1 + G6)', fontweight='bold', color='#ED7D31')
for i, v in enumerate(vd[::-1]): ax2.text(v+300, i, f'{v:,}', va='center', fontsize=8)
fig.suptitle('Diagnósticos por Tipo (Top 10) — Antes y Después', fontsize=14, fontweight='bold')
plt.tight_layout(); plt.savefig(f'{IMG_DIR}/diag.png', bbox_inches='tight'); plt.close()

# =============================================================================
# EXCEL
# =============================================================================
wb = Workbook()

# --- Hoja 1: Especialidades ---
ws1 = wb.active; ws1.title = 'Citas por Especialidad'
ws1.cell(1,1, 'Citas por Especialidad — Antes y Después del ETL').font = tf; ws1.merge_cells('A1:F1')

ws1.cell(3,1,'Especialidad'); ws1.cell(3,2,'Antes (Grupo 3)'); style_h(ws1,3,1,2,fill_a)
for i,(n,v) in enumerate(esp_antes):
    ws1.cell(4+i,1,n); ws1.cell(4+i,2,v); style_d(ws1,4+i,1,2)
r = 4+len(esp_antes)
ws1.cell(r,1,'TOTAL').font = Font(bold=True)
ws1.cell(r,2,sum(v for _,v in esp_antes)).font = Font(bold=True); ws1.cell(r,2).number_format='#,##0'; style_d(ws1,r,1,2)

ws1.cell(3,4,'Especialidad'); ws1.cell(3,5,'Después (G3+G1+G6)'); ws1.cell(3,6,'Diferencia')
style_h(ws1,3,4,6,fill_d)
ad = dict(esp_antes)
for i,(n,v) in enumerate(esp_despues):
    ws1.cell(4+i,4,n); ws1.cell(4+i,5,v)
    d = v - ad.get(n,0); ws1.cell(4+i,6,f'+{d:,}' if d>0 else str(d))
    style_d(ws1,4+i,4,6)
r = 4+len(esp_despues)
ws1.cell(r,4,'TOTAL').font = Font(bold=True)
ws1.cell(r,5,sum(v for _,v in esp_despues)).font = Font(bold=True); ws1.cell(r,5).number_format='#,##0'; style_d(ws1,r,4,6)

for c,w in [('A',22),('B',18),('D',22),('E',20),('F',14)]: ws1.column_dimensions[c].width = w
img = Image(f'{IMG_DIR}/esp.png'); img.width=900; img.height=400; ws1.add_image(img, 'A22')

# --- Hoja 2: Meses ---
ws2 = wb.create_sheet('Citas por Mes')
ws2.cell(1,1,'Citas por Mes — Antes y Después del ETL').font = tf; ws2.merge_cells('A1:E1')
ws2.cell(3,1,'Mes'); ws2.cell(3,2,'Antes (Grupo 3)'); style_h(ws2,3,1,2,fill_a)
ws2.cell(3,3,'Después (G3+G1+G6)'); ws2.cell(3,4,'Diferencia'); ws2.cell(3,5,'Incremento %')
style_h(ws2,3,3,5,fill_d)

for i,(m,a,d) in enumerate(zip(meses, mes_antes, mes_despues)):
    ws2.cell(4+i,1,m); ws2.cell(4+i,2,a); ws2.cell(4+i,3,d)
    ws2.cell(4+i,4,f'+{d-a:,}'); ws2.cell(4+i,5,f'{((d-a)/a*100):.0f}%')
    style_d(ws2,4+i,1,5)
r = 4+len(meses)
ws2.cell(r,1,'TOTAL').font = Font(bold=True)
ws2.cell(r,2,sum(mes_antes)).font = Font(bold=True); ws2.cell(r,2).number_format='#,##0'
ws2.cell(r,3,sum(mes_despues)).font = Font(bold=True); ws2.cell(r,3).number_format='#,##0'
ws2.cell(r,4,f'+{sum(mes_despues)-sum(mes_antes):,}').font = Font(bold=True)
ws2.cell(r,5,f'{((sum(mes_despues)-sum(mes_antes))/sum(mes_antes)*100):.0f}%').font = Font(bold=True)
style_d(ws2,r,1,5)
for c,w in [('A',10),('B',18),('C',20),('D',14),('E',14)]: ws2.column_dimensions[c].width = w
img2 = Image(f'{IMG_DIR}/mes.png'); img2.width=850; img2.height=380; ws2.add_image(img2, 'A19')

# --- Hoja 3: Diagnósticos ---
ws3 = wb.create_sheet('Diagnósticos por Tipo')
ws3.cell(1,1,'Diagnósticos por Tipo (Top 10) — Antes y Después del ETL').font = tf; ws3.merge_cells('A1:E1')

ws3.cell(3,1,'Tipo'); ws3.cell(3,2,'Antes (Grupo 3)'); style_h(ws3,3,1,2,fill_a)
for i,(n,v) in enumerate(diag_antes):
    ws3.cell(4+i,1,n); ws3.cell(4+i,2,v); style_d(ws3,4+i,1,2)
r = 4+len(diag_antes)
ws3.cell(r,1,'TOTAL (Top 10)').font = Font(bold=True)
ws3.cell(r,2,sum(v for _,v in diag_antes)).font = Font(bold=True); ws3.cell(r,2).number_format='#,##0'; style_d(ws3,r,1,2)

ws3.cell(3,4,'Tipo'); ws3.cell(3,5,'Después (G3+G1+G6)'); style_h(ws3,3,4,5,fill_d)
for i,(n,v) in enumerate(diag_despues):
    ws3.cell(4+i,4,n); ws3.cell(4+i,5,v); style_d(ws3,4+i,4,5)
r = 4+len(diag_despues)
ws3.cell(r,4,'TOTAL (Top 10)').font = Font(bold=True)
ws3.cell(r,5,sum(v for _,v in diag_despues)).font = Font(bold=True); ws3.cell(r,5).number_format='#,##0'; style_d(ws3,r,4,5)

for c,w in [('A',20),('B',18),('D',20),('E',20)]: ws3.column_dimensions[c].width = w
img3 = Image(f'{IMG_DIR}/diag.png'); img3.width=900; img3.height=380; ws3.add_image(img3, 'A17')

wb.save(OUTPUT)
print(f"Excel generado: {OUTPUT}")
